"""
Web routes for Yandex Maps Parser module.
"""
import os
import io
import logging
from datetime import datetime
from typing import Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, desc

from app.database import get_db
from app.models.parsed_company import ParsedCompany
from app.models.parse_task import ParseTask, ParseTaskStatus

logger = logging.getLogger(__name__)

# Setup templates
templates_path = os.path.join(os.path.dirname(__file__), "..", "web", "templates")
templates = Jinja2Templates(directory=templates_path) if os.path.exists(templates_path) else None

# Create router
router = APIRouter()


# ─── HTML Page ───────────────────────────────────────────

@router.get("/parser", response_class=HTMLResponse)
async def parser_page(request: Request):
    """Parser management page."""
    if not templates:
        return HTMLResponse("<h1>Templates not found</h1>")
    return templates.TemplateResponse("parser.html", {"request": request})


# ─── Parse Tasks API ─────────────────────────────────────

@router.get("/api/parser/tasks")
async def get_parse_tasks(db: Session = Depends(get_db)):
    """Get all parse tasks."""
    tasks = db.query(ParseTask).order_by(desc(ParseTask.created_at)).limit(100).all()
    return [t.to_dict() for t in tasks]


@router.post("/api/parser/tasks")
async def create_parse_task(data: Dict[str, Any], db: Session = Depends(get_db)):
    """Create a new parse task and launch it."""
    search_query = data.get('search_query', '').strip()
    if not search_query:
        raise HTTPException(status_code=400, detail="search_query is required")
    
    task = ParseTask(
        search_query=search_query,
        region=data.get('region', 'Москва').strip(),
        yandex_maps_url=data.get('yandex_maps_url', '').strip() or None,
        max_items=min(int(data.get('max_items', 100)), 500),
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    
    # Launch Celery task
    try:
        from parser.tasks import parse_yandex_maps_task
        result = parse_yandex_maps_task.delay(task.id)
        task.celery_task_id = result.id
        db.commit()
        logger.info(f"🚀 Parse task {task.id} launched: '{search_query}'")
    except Exception as e:
        logger.error(f"Failed to launch parse task: {e}")
        task.status = ParseTaskStatus.FAILED
        task.error_message = f"Failed to launch: {str(e)}"
        db.commit()
    
    return task.to_dict()


@router.post("/api/parser/tasks/batch")
async def create_batch_parse_tasks(data: Dict[str, Any], db: Session = Depends(get_db)):
    """Create parse tasks for multiple regions (e.g. all MO cities)."""
    search_query = data.get('search_query', '').strip()
    if not search_query:
        raise HTTPException(status_code=400, detail="search_query is required")

    regions = data.get('regions', [])
    if not regions or not isinstance(regions, list):
        raise HTTPException(status_code=400, detail="regions list is required")

    max_items = min(int(data.get('max_items', 100)), 500)
    created_tasks = []

    for region in regions:
        region = str(region).strip()
        if not region:
            continue
        task = ParseTask(
            search_query=search_query,
            region=region,
            max_items=max_items,
        )
        db.add(task)
        db.commit()
        db.refresh(task)

        try:
            from parser.tasks import parse_yandex_maps_task
            result = parse_yandex_maps_task.delay(task.id)
            task.celery_task_id = result.id
            db.commit()
        except Exception as e:
            logger.error(f"Failed to launch parse task for {region}: {e}")
            task.status = ParseTaskStatus.FAILED
            task.error_message = f"Failed to launch: {str(e)}"
            db.commit()

        created_tasks.append(task.to_dict())

    logger.info(f"🚀 Batch parse: '{search_query}' × {len(created_tasks)} cities")
    return {"created": len(created_tasks), "tasks": created_tasks}


@router.get("/api/parser/tasks/{task_id}")
async def get_parse_task(task_id: int, db: Session = Depends(get_db)):
    """Get parse task details."""
    task = db.query(ParseTask).filter(ParseTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task.to_dict()


@router.post("/api/parser/tasks/{task_id}/cancel")
async def cancel_parse_task(task_id: int, db: Session = Depends(get_db)):
    """Cancel a running parse task."""
    task = db.query(ParseTask).filter(ParseTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if task.celery_task_id:
        try:
            from tasks.celery_app import celery_app
            celery_app.control.revoke(task.celery_task_id, terminate=True)
        except Exception as e:
            logger.warning(f"Could not revoke celery task: {e}")
    
    task.status = ParseTaskStatus.CANCELLED
    task.completed_at = datetime.utcnow()
    db.commit()
    
    return {'status': 'cancelled'}


@router.delete("/api/parser/tasks/{task_id}")
async def delete_parse_task(task_id: int, db: Session = Depends(get_db)):
    """Delete a parse task (and optionally its results)."""
    task = db.query(ParseTask).filter(ParseTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    db.delete(task)
    db.commit()
    return {'status': 'deleted'}


# ─── Companies API ───────────────────────────────────────

@router.get("/api/parser/companies")
async def get_companies(
    db: Session = Depends(get_db),
    page: int = 1,
    per_page: int = 50,
    search: str = "",
    search_query: str = "",
):
    """Get parsed companies with pagination and filtering."""
    query = db.query(ParsedCompany)
    
    if search:
        query = query.filter(
            ParsedCompany.name.ilike(f"%{search}%") |
            ParsedCompany.address.ilike(f"%{search}%") |
            ParsedCompany.phone.ilike(f"%{search}%") |
            ParsedCompany.website.ilike(f"%{search}%")
        )
    
    if search_query:
        query = query.filter(ParsedCompany.search_query == search_query)
    
    total = query.count()
    companies = query.order_by(desc(ParsedCompany.created_at))\
        .offset((page - 1) * per_page)\
        .limit(per_page)\
        .all()
    
    return {
        'companies': [c.to_dict() for c in companies],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page,
    }


@router.delete("/api/parser/companies/{company_id}")
async def delete_company(company_id: int, db: Session = Depends(get_db)):
    """Delete a parsed company."""
    company = db.query(ParsedCompany).filter(ParsedCompany.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    db.delete(company)
    db.commit()
    return {'status': 'deleted'}


@router.delete("/api/parser/companies")
async def delete_all_companies(db: Session = Depends(get_db), search_query: str = ""):
    """Delete all companies (optionally filtered by search_query)."""
    query = db.query(ParsedCompany)
    if search_query:
        query = query.filter(ParsedCompany.search_query == search_query)
    
    count = query.count()
    query.delete(synchronize_session=False)
    db.commit()
    
    return {'status': 'deleted', 'count': count}


# ─── Stats ───────────────────────────────────────────────

@router.get("/api/parser/stats")
async def get_parser_stats(db: Session = Depends(get_db)):
    """Get parser statistics."""
    total_companies = db.query(func.count(ParsedCompany.id)).scalar() or 0
    total_tasks = db.query(func.count(ParseTask.id)).scalar() or 0
    running_tasks = db.query(func.count(ParseTask.id)).filter(
        ParseTask.status == ParseTaskStatus.RUNNING
    ).scalar() or 0
    
    # Unique search queries
    unique_queries = db.query(func.count(func.distinct(ParsedCompany.search_query))).scalar() or 0
    
    # Companies with contacts
    with_phone = db.query(func.count(ParsedCompany.id)).filter(
        ParsedCompany.phone.isnot(None)
    ).scalar() or 0
    with_email = db.query(func.count(ParsedCompany.id)).filter(
        ParsedCompany.email.isnot(None)
    ).scalar() or 0
    with_website = db.query(func.count(ParsedCompany.id)).filter(
        ParsedCompany.website.isnot(None)
    ).scalar() or 0
    
    return {
        'total_companies': total_companies,
        'total_tasks': total_tasks,
        'running_tasks': running_tasks,
        'unique_queries': unique_queries,
        'with_phone': with_phone,
        'with_email': with_email,
        'with_website': with_website,
    }


# ─── Email Extraction ───────────────────────────────────

@router.post("/api/parser/extract-emails")
async def extract_emails(data: Dict[str, Any] = {}, db: Session = Depends(get_db)):
    """Launch email extraction task for companies without email."""
    from parser.tasks import extract_emails_task

    search_query = data.get('search_query', None)

    # Count how many need email
    query = db.query(func.count(ParsedCompany.id)).filter(
        ParsedCompany.website.isnot(None),
        ParsedCompany.website != '',
        (ParsedCompany.email.is_(None)) | (ParsedCompany.email == ''),
    )
    if search_query:
        query = query.filter(ParsedCompany.search_query == search_query)
    need_email = query.scalar() or 0

    if need_email == 0:
        return {'status': 'nothing_to_do', 'message': 'Все компании уже имеют email или не имеют сайта'}

    task = extract_emails_task.apply_async(
        kwargs={'search_query': search_query},
        queue='parser',
    )

    return {
        'status': 'started',
        'celery_task_id': task.id,
        'companies_to_process': need_email,
        'message': f'Запущено извлечение email для {need_email} компаний'
    }


# ─── Export XLS ──────────────────────────────────────────

@router.get("/api/parser/export")
async def export_companies_xls(
    db: Session = Depends(get_db),
    search_query: str = "",
):
    """Export companies to XLSX file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    query = db.query(ParsedCompany)
    if search_query:
        query = query.filter(ParsedCompany.search_query == search_query)
    
    companies = query.order_by(ParsedCompany.id).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Компании"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "№", "Название", "Категория", "Адрес", "Телефон", "Телефон 2",
        "Email", "Сайт", "Telegram", "WhatsApp", "VK", "Instagram",
        "Рейтинг", "Отзывов", "Часы работы", "Запрос", "Регион",
        "Ссылка Яндекс.Карты"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, company in enumerate(companies, 2):
        values = [
            row_idx - 1,
            company.name,
            company.category,
            company.address,
            company.phone,
            company.phone2,
            company.email,
            company.website,
            company.telegram,
            company.whatsapp,
            company.vk,
            company.instagram,
            company.rating,
            company.reviews_count,
            company.working_hours,
            company.search_query,
            company.region,
            company.yandex_maps_url,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    
    filename = f"companies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    from fastapi.responses import Response
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(content)),
        }
    )
